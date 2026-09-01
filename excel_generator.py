import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_surat_tugas_excel(records, sheet_title="Surat Tugas"):
    """
    Generates a professionally styled Excel workbook for HR Surat Tugas matching MBA ITB template.
    Column F (Teaching Date) is hidden, Transportation is removed.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] # Excel max sheet title is 31 chars
    
    # Styles
    header_font = Font(name="Tahoma", size=9, bold=True, color="1F2937")
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    data_font = Font(name="Calibri", size=11, color="000000")
    tahoma_font = Font(name="Tahoma", size=9, color="000000")
    
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Header Row
    headers = [
        "No.", 
        "Name", 
        "Status", 
        "Program", 
        "Course", 
        "Date", 
        "Date (Tanggal Surat Tugas)", 
        "Ditugaskan Sebagai", 
        "Lokasi"
    ]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    
    # Data Rows
    for idx, r in enumerate(records, start=1):
        row_num = idx + 1
        ws.append([
            idx,
            r.get("matched_name") or r.get("lecturer", ""),
            r.get("status", ""),
            r.get("program", ""),
            r.get("course", ""),
            r.get("iso_date", ""),
            r.get("tanggal_surat_tugas", ""),
            r.get("ditugaskan_sebagai", "Dosen Pengajar"),
            r.get("lokasi", "SBM ITB Kampus Jakarta")
        ])
        
        # Apply formatting to data cells
        # 1: No
        ws.cell(row=row_num, column=1).font = tahoma_font
        ws.cell(row=row_num, column=1).alignment = align_center
        ws.cell(row=row_num, column=1).border = thin_border
        
        # 2: Name
        ws.cell(row=row_num, column=2).font = data_font
        ws.cell(row=row_num, column=2).alignment = align_left
        ws.cell(row=row_num, column=2).border = thin_border
        
        # 3: Status
        ws.cell(row=row_num, column=3).font = data_font
        ws.cell(row=row_num, column=3).alignment = align_center
        ws.cell(row=row_num, column=3).border = thin_border
        
        # 4: Program
        ws.cell(row=row_num, column=4).font = data_font
        ws.cell(row=row_num, column=4).alignment = align_center
        ws.cell(row=row_num, column=4).border = thin_border
        
        # 5: Course
        ws.cell(row=row_num, column=5).font = data_font
        ws.cell(row=row_num, column=5).alignment = align_left
        ws.cell(row=row_num, column=5).border = thin_border
        
        # 6: Date Mengajar (Hidden)
        ws.cell(row=row_num, column=6).font = data_font
        ws.cell(row=row_num, column=6).alignment = align_center
        ws.cell(row=row_num, column=6).border = thin_border
        
        # 7: Tanggal Surat Tugas
        ws.cell(row=row_num, column=7).font = tahoma_font
        ws.cell(row=row_num, column=7).alignment = align_center
        ws.cell(row=row_num, column=7).border = thin_border
        
        # 8: Ditugaskan Sebagai
        ws.cell(row=row_num, column=8).font = data_font
        ws.cell(row=row_num, column=8).alignment = align_center
        ws.cell(row=row_num, column=8).border = thin_border
        
        # 9: Lokasi
        ws.cell(row=row_num, column=9).font = data_font
        ws.cell(row=row_num, column=9).alignment = align_center
        ws.cell(row=row_num, column=9).border = thin_border
        
    # Column Dimensions
    ws.column_dimensions["A"].width = 7.0
    ws.column_dimensions["B"].width = 48.0
    ws.column_dimensions["C"].width = 14.0
    ws.column_dimensions["D"].width = 15.0
    ws.column_dimensions["E"].width = 65.0
    ws.column_dimensions["F"].width = 16.0
    ws.column_dimensions["F"].hidden = True # Hidden teaching date (Kolom F)
    ws.column_dimensions["G"].width = 25.0
    ws.column_dimensions["H"].width = 22.0
    ws.column_dimensions["I"].width = 25.0
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
